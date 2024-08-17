import flet as ft
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
from time import sleep

workbook=Workbook()
worksheet=workbook.active

def show_files(e):
        column.controls.clear()
        for root, dirs, files in os.walk(".\\files"):
            for name in files:
                if ".xlsx" in name or ".XLSX" in name or ".xls" in name:
                  column.controls.append(
                       ft.Container(
                       bgcolor="#96B6C5",
                       content=ft.Row([ft.Text(f"{name}", size=18)], alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                  ))
                  )
                  files_list.append(os.path.join(root, name))
            else:
                  column.controls.append(ft.Text("Above are the only files found", color=ft.colors.WHITE, size=18))  
            container.update()

def create_xl_from_df(save_name, voucher_list):
     df=pd.DataFrame(voucher_list, columns=col)
     df_sorted = df.sort_values(by=['Date'], ascending=[True])
     df_sorted.to_excel(f"{save_name.replace(".\\files", ".\\output")}",index=False)
    #  df.to_excel(f"{save_name.replace(".\\files", ".\\output")}",index=False)
    
def format_excel(e):
     for file in files_list:
        df = pd.read_excel(file,skiprows=3)
        df.to_excel(file,index=False)
        wb=load_workbook(file)
        ws=wb.active
        list_for_df=[]
        for row in ws.iter_rows():
             # add if statement for local and center
             if "2B" in file:
              if "23" in row[2].value:
                item_type="L/GST-ItemWise" if sale_purchase_type.value=="Old version" else "Local-ItemWise"

              else: 
                item_type="I/GST-ItemWise" if sale_purchase_type.value=="Old version" else "Center-ItemWise" 
              
              nested_list=["main",row[1].value,row[5].value,row[7].value,row[9].value,row[10].value, item_type]
             
             else:
                if "23" in row[2].value:
                  item_type="L/GST-ItemWise" if sale_purchase_type.value=="Old version" else "Local-ItemWise"

                else: 
                  item_type="I/GST-ItemWise" if sale_purchase_type.value=="Old version" else "Center-ItemWise"
                nested_list=["main",row[1].value,row[5].value,row[8].value,row[10].value,row[11].value, item_type]
             list_for_df.append(nested_list)
        create_xl_from_df(file,list_for_df)
        dlg = ft.AlertDialog(
                    title=ft.Text(f"{file} has been formated")
        )
        e.page.dialog = dlg
        dlg.open = True
        e.page.update()
        sleep(2)
        dlg.open = False
        e.page.update()
        wb.close()

col=["Series", "Party Name", "Invoice No.", "Date", "Taxable Amount", "Item Name", "Sale/Purchase Type"]

# list_for_df=[]

files_list=[]

column=ft.Column()
    
refresh_btn=ft.IconButton(icon=ft.icons.REFRESH, 
                              icon_color=ft.colors.BLUE, tooltip="Refresh", on_click=show_files)

sale_purchase_type=ft.Dropdown(
     height=45,
     width=150,
     content_padding=0,
     filled=True,
     alignment=ft.alignment.center,
     bgcolor="#252B48",
     color=ft.colors.WHITE,
     options=[
          ft.dropdown.Option("Old version"),
          ft.dropdown.Option("New version")
     ]
)

start_btn=ft.ElevatedButton(text="Start",
                                 icon_color=ft.colors.BLUE, tooltip="Refresh", on_click=format_excel)

back_btn=ft.IconButton(icon=ft.icons.U_TURN_LEFT_SHARP, 
                              icon_color=ft.colors.RED, tooltip="Back", on_click=lambda e: e.page.go("/index"))

container=ft.Container(
                        border=ft.border.all(5, ft.colors.PINK_600),
                        bgcolor="#252B48",
                        height=300,
                        width=750,
                        tooltip="Data Table",
                        content=column
                    )

def __view__():
    return ft.View(
        "/excel",
        controls=[
            ft.Column(
                [   
                    ft.Row([container],alignment=ft.MainAxisAlignment.CENTER),
                    ft.Row([refresh_btn, 
                            sale_purchase_type,
                            start_btn,
                            back_btn
                            ],
                            alignment=ft.MainAxisAlignment.CENTER)
                ], alignment=ft.MainAxisAlignment.CENTER
            )
        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        vertical_alignment=ft.MainAxisAlignment.CENTER
    )